import os
import argparse
import openpyxl
from openpyxl.styles import Alignment, PatternFill, Border, Side

parser = argparse.ArgumentParser()
parser.add_argument('--in_path', type = str, default = './', help = 'path to the stats file')
parser.add_argument('--out_path', type = str, default = './', help = 'path to the output file')
parser.add_argument('--in_file', type = str, default = 'stats', help = 'name of the stats file')
parser.add_argument('--out_file', type = str, default = 'stats', help = 'name of the output spreadsheet file')
parser.add_argument('--sheet_name', type = str, default = 'Sheet', help = 'name of this spreadsheet')

args = parser.parse_args()

in_path = str(args.in_path)
out_path = str(args.out_path)
in_file = str(args.in_file)
out_file = str(args.out_file)
sheet_name = str(args.sheet_name)
col_no = 1
thin = Side(border_style="thin", color="000000")

#convert column number to alphabetical letters
def colnum_string(n):
    string = ""
    while n > 0:
        n, remainder = divmod(n - 1, 26)
        string = chr(65 + remainder) + string
    return string

#get the value of corresponding pattern from stats file
def getMatchingValue(pattern):
	stats_file.seek(0)
	lines = stats_file.readlines()
	for line in lines:
		if pattern in line:
			words = list(line.split(" "))
			return int(words[len(words)-1])
	return 0;

#fill the colors into specified portion of excel sheet
def fillCellColor(row_start, col_start, row_end, col_end, color):
	for row in range(row_start, row_end+1):
		for col in range(col_start, col_end+1):
			worksheet[colnum_string(col)+str(row)].fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

#fill each column of excel sheet
def fillColumn(heading, start_pattern, end_pattern, row_no = 11, rows = 38, factor = 1):
	global col_no
	worksheet.column_dimensions[colnum_string(col_no)].width = 15
	worksheet[colnum_string(col_no)+str(row_no)] = heading
	worksheet[colnum_string(col_no)+str(row_no)].border = Border(top=thin, left=thin, right=thin, bottom=thin)
	worksheet[colnum_string(col_no)+str(row_no)].alignment = Alignment(horizontal='justify')
	sum = 0
	for r in range(rows):
		row_no += 1
		value = getMatchingValue(start_pattern+str(r)+end_pattern) * factor
		sum += value
		worksheet[colnum_string(col_no)+str(row_no)].border = Border(top=thin, left=thin, right=thin, bottom=thin)
		worksheet[colnum_string(col_no)+str(row_no)] = value
	worksheet[colnum_string(col_no)+str(row_no+1)] = sum
	col_no += 1
	
#open statistics file	
stats_file = open(in_path+'/'+in_file+'.stats')


# Create an new Excel file and add a worksheet.
os.system('mkdir -p ' + out_path)

if os.path.exists(out_path+'/'+out_file+'.xlsx'):
	workbook = openpyxl.load_workbook(out_path+'/'+out_file+'.xlsx')
	worksheet = workbook.create_sheet(title = sheet_name)
else:
	workbook = openpyxl.Workbook()
	worksheet = workbook.active
	worksheet.title = sheet_name

#print left side labels in excel sheet
for core in range(38):
	worksheet['A'+str(core+12)] = 'Core'+str(core)

worksheet.column_dimensions['A'].width = 12
#increase height of the Sub-Heading Row
worksheet.row_dimensions[11].height = 70

#print the Heading
def printHeading(name, col, row):
    worksheet[colnum_string(col)+str(row)] = name
    
#set style of Heading (merge cells and align to center)
def setStyle(start, end, color):
    worksheet.merge_cells(colnum_string(start)+'10:'+colnum_string(end)+'10')
    worksheet[colnum_string(start)+'10'].alignment = Alignment(horizontal='center')
    fillCellColor(11,start, 49,end, color)


#***************************MLC*****************************************#
col_no += 1
printHeading('MLC', col_no, 10)
start = col_no

fillColumn('Tot. command (uncore)','SBR.p0.cache_level_0_','_cluster_CACHE.instance_0_total_read_command',factor = 0.5)
fillColumn('Hit (uncore)','SBR.p0.cache_level_0_','_cluster_CACHE.instance_0_hit',factor = 0.5)
fillColumn('Miss (uncore cmnd)','SBR.p0.cache_level_0_','_cluster_CACHE.instance_0_miss',factor = 0.5)
fillColumn('Eviction (Capacity)','SBR.p0.cache_level_0_','_cluster_CACHE.instance_0_capacity_evictions')
fillColumn('SF_BackInv (Hit in MLC)','SBR.p0.cache_level_0_','_cluster_CACHE.instance_0_sfceviction')
fillColumn('share','SBR.p0.cache_level_0_','_cluster_CACHE.instance_0_total_share')
fillColumn('HSF_BackInv (ESnpInv) (Hit in MLC)','SBR.p0.cache_level_0_','_cluster_CACHE.instance_0_HIT_esnpinv')
fillColumn('CRdD', 'SBR.p0.cache_level_0_' , '_cluster_CACHE.instance_0_IDI_Opc_CRdD_count')
fillColumn('DRdD', 'SBR.p0.cache_level_0_' , '_cluster_CACHE.instance_0_IDI_Opc_DRdD_count')
fillColumn('RFO', 'SBR.p0.cache_level_0_' , '_cluster_CACHE.instance_0_IDI_Opc_RFO_count')


setStyle(start, col_no-1, '00FFFF99')



#****************************SF*****************************************#
col_no += 1
printHeading('SF', col_no, 10)
start = col_no

fillColumn('Tot. command','SBR.p0.cache_level_1_0_cluster_SF.instance_','_total_read_command',factor = 0.5)
fillColumn('Hit (uncore)','SBR.p0.cache_level_1_0_cluster_SF.instance_','_hit',factor = 0.5)
fillColumn('Miss; M = LLC_Hits+HSF_reads','SBR.p0.cache_level_1_0_cluster_SF.instance_','_miss',factor = 0.5)
fillColumn('Capacity_Evcition','SBR.p0.cache_level_1_0_cluster_SF.instance_','_capacity_evictions')
fillColumn('SF_Binv(Shared)','SBR.p0.cache_level_1_0_cluster_SF.instance_','_sfbackinv_sharedblock')
fillColumn('SF_Binv(Exclusive)','SBR.p0.cache_level_1_0_cluster_SF.instance_','_sfbackinv_exclusiveblock')
fillColumn('share','SBR.p0.cache_level_1_0_cluster_SF.instance_','_total_share')
fillColumn('Evictions from MLC resulting into SF eviction(WbEtoI+WbStoI+WbMtoI)','SBR.p0.cache_level_1_0_cluster_SF.instance_','_writeback')
fillColumn('Invalidation in SF(due to MLC Eviction)','SBR.p0.cache_level_1_0_cluster_SF.instance_','_invalidate')
fillColumn('HSF_BackInv (hit in SF)','SBR.p0.cache_level_1_0_cluster_SF.instance_','_sfceviction')
fillColumn('SF_generated BackInv','SBR.p0.cache_level_1_0_cluster_SF.instance_','_MISS_backinv',factor = 0.5)
fillColumn('Snoops','SBR.p0.cache_level_1_0_cluster_SF.instance_','_Snoops' )
fillColumn('CRdD', 'SBR.p0.cache_level_1_0_cluster_SF.instance_' , '_IDI_Opc_CRdD_count')
fillColumn('DRdD', 'SBR.p0.cache_level_1_0_cluster_SF.instance_' , '_IDI_Opc_DRdD_count')
fillColumn('RFO', 'SBR.p0.cache_level_1_0_cluster_SF.instance_' , '_IDI_Opc_RFO_count')

setStyle(start, col_no-1, '0099CCFF')

#***************************LLC*****************************************#
col_no += 1
printHeading('LLC', col_no, 10)
start = col_no

fillColumn('Tot. command','SBR.p0.cache_level_1_0_cluster_SF.instance_','_total_read_command',factor = 0.5)
fillColumn('Hit (Data hit)','SBR.p0.cache_level_1_0_cluster_CACHE.instance_','_hit',factor = 0.5)
fillColumn('Miss; M = M-W','SBR.p0.cache_level_1_0_cluster_CACHE.instance_','_miss',factor = 0.5)
fillColumn('LLC capicity eviction','SBR.p0.cache_level_1_0_cluster_CACHE.instance_','_capacity_evictions')
fillColumn('HSF_generated_BackInv (hit in llc)','SBR.p0.cache_level_1_0_cluster_CACHE.instance_','_sfceviction')
fillColumn('LLC generated WB','SBR.p0.cache_level_1_0_cluster_CACHE.instance_','_IDI_Opc_Writeback_lookups',factor = 0.5)
fillColumn('INVALIDATE = LLC to SF migration','SBR.p0.cache_level_1_0_cluster_CACHE.instance_','_invalidate')


setStyle(start, col_no-1, '00CCFFCC')

#***************************HSF*****************************************#
col_no += 1
printHeading('HSF', col_no, 10)
start = col_no

fillColumn('Total Command = Cache miss','SBR.p0.cache_level_2_0_cluster_SF.instance_','_total_read_command',factor = 0.5)
fillColumn('Mis','SBR.p0.cache_level_2_0_cluster_SF.instance_','_miss',factor = 0.5)
fillColumn('Snoops','SBR.p0.cache_level_2_0_cluster_SF.instance_','_Snoops')
fillColumn('Evictions(Capacity)','SBR.p0.cache_level_2_0_cluster_SF.instance_','_capacity_evictions')
fillColumn('Page eviction','SBR.p0.cache_level_2_0_cluster_SF.instance_','_page_evictions')
fillColumn('Page hit','SBR.p0.cache_level_2_0_cluster_SF.instance_','_page_hit')
fillColumn('Clean Evicts from LLC rerequested by core later','SBR.p0.cache_level_2_0_cluster_SF.instance_','_invalidated_page_hit', factor=0.5)
fillColumn('LLC generated WB (hit in HSF)','SBR.p0.cache_level_2_0_cluster_SF.instance_','_LLC_writeback')
fillColumn('CRdD', 'SBR.p0.cache_level_2_0_cluster_SF.instance_' , '_IDI_Opc_CRdD_count')
fillColumn('DRdD', 'SBR.p0.cache_level_2_0_cluster_SF.instance_' , '_IDI_Opc_DRdD_count')
fillColumn('RFO', 'SBR.p0.cache_level_2_0_cluster_SF.instance_' , '_IDI_Opc_RFO_count')

setStyle(start, col_no-1, '00FFCC99')

#******************************Top Portion******************************#
worksheet['A1'] = 'S_x = S_0'
worksheet['A5'] = 'Cache_L_0'
worksheet['A6'] = 'Cache_L_1'
worksheet['A7'] = 'Cache_L_2'

worksheet['B1'] = 'Total command'
worksheet['B2'] = '=V2'
worksheet['B3'] = '=B2-B50'
worksheet['B4'] = 'Hits(Uncore)'
worksheet['B5'] = '=C50'
worksheet['B6'] = '=N50+AD50' #check it once for the correct columns
worksheet['B7'] = 0

worksheet['C4'] = 'Miss(Uncore)'
worksheet['C5'] = '=D50'
worksheet['C6'] = '=O50+AE50' #check it once for the correct columns
worksheet['C7'] = '=AL50' #check it once for the correct columns

worksheet['D4'] = 'Total'
worksheet['D5'] = '=B5+C5'
worksheet['D6'] = '=C5-B6' #check it once for the correct columns

worksheet['E6'] = 'SF+CACHE'

worksheet['J1'] = 'Reads'
worksheet['J2'] = '=V2'
worksheet['J4'] = 'Snoops'
worksheet['J5'] = 0
worksheet['J6'] = '=X50' #check it once for the correct columns
worksheet['J7'] = '=AM50' #check it once for the correct columns

worksheet['K1'] = 'Writes (Clean)'
worksheet['K2'] = 0
worksheet['L1'] = 'Writes (Dirty)'
worksheet['L2'] = 0

worksheet['S1'] = 'CRdD'
worksheet['T1'] = 'DRdD'
worksheet['U1'] = 'RFO'
worksheet['V1'] = 'Total'
worksheet['V2'] = '=S2+T2+U2'

fillCellColor(1,1, 2,22, '00FF9900')

#save and close the workbook
workbook.save(out_path+'/'+out_file+'.xlsx')
workbook.close()

#close the stats file
stats_file.close()
