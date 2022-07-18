# -*- coding: utf-8 -*-
"""
Created on Sat Jun 18 19:10:31 2022

@author: bbisht
"""

import argparse
import openpyxl
from openpyxl.styles import Alignment, PatternFill, Border, Side, Font
from openpyxl.styles.numbers import FORMAT_PERCENTAGE

parser = argparse.ArgumentParser()
parser.add_argument('--in_path', type = str, default = './', help = 'path to the stats file')
parser.add_argument('--in_file', type = str, default = 'stats', help = 'name of the stats file')
parser.add_argument('--sheet_name', type = str, default = 'Summary', help = 'name of this spreadsheet')

args = parser.parse_args()

in_path = str(args.in_path)
in_file = str(args.in_file)
sheet_name = str(args.sheet_name)
thin = Side(border_style="thin", color="000000")
head_font = Font(size=13, bold=True)
subhead_font = Font(size=13)
benchmarks = ['astar1', 'astarHMPKI', 'graph500', 'omnetpp', 'stream', 'xsbench']
sectors = ['512']

#convert column number to alphabetical letters
def colnum_string(n):
    string = ""
    while n > 0:
        n, remainder = divmod(n - 1, 26)
        string = chr(65 + remainder) + string
    return string






#fill the colors into specified portion of excel sheet
def fillCellColor(row_start, col_start, row_end, col_end, color):
	for row in range(row_start, row_end+1):
		for col in range(col_start, col_end+1):
			worksheet[colnum_string(col)+str(row)].fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    
    
#print the Heading
def printHeading(name, col, row):
    worksheet[colnum_string(col)+str(row)] = name   
    worksheet[colnum_string(col)+str(row)].font = head_font
    
def printSubHeading(name, col, row):
    worksheet[colnum_string(col)+str(row)] = name   
    worksheet[colnum_string(col)+str(row)].font = subhead_font
    


# open results excel-file
workbook = openpyxl.load_workbook(in_path+'/'+in_file+'.xlsx')
worksheet = workbook.create_sheet(title = sheet_name)

#print left side labels in excel sheet
worksheet.column_dimensions['A'].width = 30
worksheet.row_dimensions[1].height = 30
printHeading('Baseline with Trace command injections', 1, 1)
printHeading('HSF Size', 1, 2)
printSubHeading('MLC Miss %', 1, 3)
worksheet['A4'] = 'LLC Hit %'
worksheet['A5'] = 'Core-to-core transfer %'
worksheet['A6'] = 'Data from DRAM %'
printSubHeading('SF Snoops breakdown %', 1, 7)
worksheet['A8'] = 'shared data %'
worksheet['A9'] = 'remaining (migrate) %'
worksheet['A10'] = 'SF BackInv % (eviction%)'
worksheet['A11'] = 'SF data BackInv %'
printSubHeading('HSF miss (adds to DRAM%)', 1, 12)
worksheet['A13'] = 'HSF sector miss %'
worksheet['A14'] = 'HSF False hit %'
worksheet['A15'] = 'HSF BackInv %'
worksheet['A16'] = 'HSF BackInv to SF %'
worksheet['A17'] = 'HSF BackInv to LLC %'
worksheet['A18'] = 'avg CL\'s BackInv per sector'



col_no = 2
for bench in benchmarks:
    printHeading(bench, col_no, 1)
    worksheet.merge_cells(colnum_string(col_no)+'1:'+colnum_string(col_no+len(sectors)-1)+'1')
    worksheet[colnum_string(col_no)+'1'].alignment = Alignment(horizontal='center')
    
    worksheet[colnum_string(col_no)+'3'] = '='+bench+'_32MC_'+sectors[0]+'!D50/'+bench+'_32MC_'+sectors[0]+'!B50' 
    worksheet[colnum_string(col_no)+'4'] = '='+bench+'_32MC_'+sectors[0]+'!AD50/'+bench+'_32MC_'+sectors[0]+'!B50'
    worksheet[colnum_string(col_no)+'5'] = '='+bench+'_32MC_'+sectors[0]+'!N50/'+bench+'_32MC_'+sectors[0]+'!B50'    
    worksheet[colnum_string(col_no)+'6'] = '='+bench+'_32MC_'+sectors[0]+'!AK50/'+bench+'_32MC_'+sectors[0]+'!B50'    
    worksheet[colnum_string(col_no)+'7'] = '='+colnum_string(col_no)+'5'    
    worksheet[colnum_string(col_no)+'8'] = '='+bench+'_32MC_'+sectors[0]+'!S50/'+bench+'_32MC_'+sectors[0]+'!B50'    
    worksheet[colnum_string(col_no)+'9'] = '=('+bench+'_32MC_'+sectors[0]+'!N50-'+bench+'_32MC_'+sectors[0]+'!S50)/'+bench+'_32MC_'+sectors[0]+'!B50'
    worksheet[colnum_string(col_no)+'10'] = '='+bench+'_32MC_'+sectors[0]+'!P50/'+bench+'_32MC_'+sectors[0]+'!B50'
    worksheet[colnum_string(col_no)+'11'] = '='+bench+'_32MC_'+sectors[0]+'!F50/'+bench+'_32MC_'+sectors[0]+'!B50'
    worksheet[colnum_string(col_no)+'12'] = '='+colnum_string(col_no)+'6'
    
    for row in range(3,13):
        worksheet[colnum_string(col_no)+str(row)].number_format = FORMAT_PERCENTAGE
        worksheet[colnum_string(col_no)+str(row)].alignment = Alignment(horizontal='center')
        worksheet.merge_cells(colnum_string(col_no)+str(row)+':'+colnum_string(col_no+len(sectors)-1)+str(row))
        
    for sector in sectors:
        worksheet.column_dimensions[colnum_string(col_no)].width = 13
        printSubHeading(sector+'B', col_no, 2)
        worksheet[colnum_string(col_no)+'13'] = '=('+bench+'_32MC_'+sector+'!AK50-'+bench+'_32MC_'+sector+'!AP50)/'+bench+'_32MC_'+sector+'!B50'
        worksheet[colnum_string(col_no)+'14'] = '='+bench+'_32MC_'+sector+'!AP50/'+bench+'_32MC_'+sector+'!B50'
        worksheet[colnum_string(col_no)+'15'] = '='+bench+'_32MC_'+sector+'!AM50/'+bench+'_32MC_'+sector+'!B50'
        worksheet[colnum_string(col_no)+'16'] = '='+bench+'_32MC_'+sector+'!H50/'+bench+'_32MC_'+sector+'!B50'
        worksheet[colnum_string(col_no)+'17'] = '='+bench+'_32MC_'+sector+'!AG50/'+bench+'_32MC_'+sector+'!B50'
        
        worksheet[colnum_string(col_no)+'18'] = '='+bench+'_32MC_'+sector+'!AM50/'+bench+'_32MC_'+sector+'!AN50/100'
        for row in range(13,19):
            worksheet[colnum_string(col_no)+str(row)].number_format = FORMAT_PERCENTAGE
            worksheet[colnum_string(col_no)+str(row)].alignment = Alignment(horizontal='center')
        col_no += 1


printHeading('Commnets', col_no, 1)
worksheet.column_dimensions[colnum_string(col_no)].width = 20
worksheet.merge_cells(colnum_string(col_no)+'1:'+colnum_string(col_no)+'2')
fillCellColor(1, 1, 2, col_no, '00CCCCFF')


#save and close the workbook
workbook.save(in_path+'/'+in_file+'.xlsx')
workbook.close()
